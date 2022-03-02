#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jan 20 12:10:24 2021

@author: aletorella
"""
import xlwings as xw
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
sns.set()
import datetime as datetime
from timeit import default_timer as timer
from openpyxl import load_workbook

# file_names=['AL','AO','AT','BI','CN','IM','NO','SV','VB','VC']
posizioniAL = np.array([14,26,38,50,62,74,86,98,110])
posizioniAO = np.array([14,26,38,50,62,74,86,98])
posizioniAT = np.array([14,26,38,50,62,74])
posizioniBI = np.array([14,26,38,50,62,74,86])
posizioniCN = np.array([14,26,38,50,62,74,86,98])
posizioniIM = np.array([14,26,38,50,62,74,86,98])
posizioniNO = np.array([14,26,38,50,62,74,86,98])
posizioniSV = np.array([14,26,38,50,62,74,86])
posizioniVB = np.array([14,26,38,50,62,74,86])
posizioniVC = np.array([14,26,38,50,62,74,86,98])

def colonne(df):
    col_list = [ chr(i+65) for i in range(len(df.columns)) ]
    df.columns = col_list
    return df

####################FUNZIONE CHE SCARICA FILE DIVISI PER TESTATA#########################
########Questa funzione salva il file di ogni testata separato per regione#############
##########################(da inserire manualmente)##############inserire regione come stringa#####
def get_segB(regione,posizioni,mese):
        start = timer()
        month = {'1':11,'2':10,'3':9,'4':8,'5':7,'6':6,'7':5,'8':4,'9':3,'10':2,'11':1,'12':0}
        righe = {'1':8,'2':9,'3':10,'4':11,'5':12,'6':13,'7':14,'8':15,'9':16,'10':17,'11':18,'12':19}
        pos = posizioni-month[str(mese)]
        
        
        df = pd.read_excel('DB CONCORRENZA SEG. B.xlsx',regione)
        col_list = [ chr(i+65) for i in range(len(df.columns)) ]
        df.columns = col_list
        
        
        new_file = pd.ExcelFile('SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx')#salva temporaneamente il file completo
        lista_fogli = new_file.sheet_names#prende il nome delle testate 
        k=0
        wb = load_workbook('SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx')
        
        for j in range(1,len(lista_fogli)):
            wb.active = j
            ws = wb.active
            testata = lista_fogli[j]
            
            vars()[lista_fogli[j]] = pd.read_excel('SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx', testata)
            file_testata = vars()[lista_fogli[j]]
            col_list = [ chr(i+65) for i in range(len(file_testata.columns)) ]
            
            
            file_testata.columns = col_list
            pos_db = pos[k]
            valore1 = df.at[pos_db,'E'] #prendo il SOMMA DI VENDUTO
            valore2 = df.at[pos_db,'M'] #prendo il MEDIA COPIE GIORNO
            
            ws.cell(row=righe[str(mese)], column=5).value = valore1
            ws.cell(row=righe[str(mese)], column=9).value = valore2
            
            wb.save(filename='SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx')
            k+=1
            
        end = timer()
        return end-start
    
    
def file_lastampa():
    start = timer()
    lettere=['B','C','D','E','F','G','H']
    regioni = ['AL','AO','AT','BI','CN','IM','NO','SV','VB','VC']
    stampa = pd.read_excel('PER LA STAMPA SETT_NOV 2020.xlsx','NOVEMBRE 2020')
    stampa = colonne(stampa)
    k=17
    k1=3
    for i in regioni:
        wb = load_workbook('SEGMENTO B_Concorrenza Gen17-Nov20_'+i+'.xlsx')
        wb.active = 0
        ws = wb.active
        c=3
        for j in lettere:
            value1 = stampa.at[k,j]
            value2 = stampa.at[k1,'I']
            ws.cell(row = 8, column = c).value = value1
            ws.cell(row = 8, column = 12).value = value2
            c+=1
        wb.save(filename='SEGMENTO B_Concorrenza Gen17-Nov20_'+i+'.xlsx')
        k += 1
        k1 += 1
    end = timer()
    return end-start
        
        
    
    
    
    

# wb = load_workbook('PER LA STAMPA SETT_NOV 2020.xlsx', data_only=True)
# ws = wb.active
# stampa = pd.read_excel('PER LA STAMPA SETT_NOV 2020.xlsx','NOVEMBRE 2020')
# stampa = colonne(stampa)

wb = load_workbook('SEGMENTO B_Concorrenza Gen17-Nov20_AL.xlsx')
   
    
    
    

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
# def get_segB(regione,posizioni,mese):
#         start = timer()
#         month = {'01':11,'02':10,'03':9,'04':8,'05':7,'06':6,'07':5,'08':4,'09':3,'10':2,'11':1,'12':0}
#         pos = posizioni-month[str(mese)]
        
        
#         df = pd.read_excel('DB CONCORRENZA SEG. B.xlsx',regione)
#         col_list = [ chr(i+65) for i in range(len(df.columns)) ]
#         df.columns = col_list
        
        
#         new_file = pd.ExcelFile('SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx')#salva temporaneamente il file completo
#         lista_fogli = new_file.sheet_names #prende il nome delle testate 
#         k=0
#         for j in lista_fogli[1:]:
#             testata = j
            
#             vars()[j] = pd.read_excel('SEGMENTO B_Concorrenza Gen17-Nov20_'+regione+'.xlsx', testata)
#             file_testata = vars()[j]
#             col_list = [ chr(i+65) for i in range(len(file_testata.columns)) ]
            
            
#             file_testata.columns = col_list
#             pos_db = pos[k]
#             valore1 = df.at[pos_db,'E'] #prendo il SOMMA DI VENDUTO
#             valore2 = df.at[pos_db,'M'] #prendo il MEDIA COPIE GIORNO
#             file_testata.at[17,'E'] = valore1
#             file_testata.at[17,'I'] = valore2
            
            
#             file_testata.to_excel(str(testata)+'_'+regione+'_'+str(mese)+'.xlsx')
#             k+=1
            
#         end = timer()
#         return end-start


            
                
            
    
    
    
    
        
        
            
        
            
            
        
    
    
    
    
    
    
    
    






